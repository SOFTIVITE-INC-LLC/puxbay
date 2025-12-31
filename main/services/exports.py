"""
Export Service for generating CSV and Excel reports.
Provides utilities for exporting various data types from the POS system.
"""
import csv
import io
from datetime import datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def export_to_csv(queryset, fields, filename='export'):
    """
    Export a queryset to CSV format.
    
    Args:
        queryset: Django queryset to export
        fields: List of tuples (field_name, display_name)
        filename: Base filename for the export
    
    Returns:
        HttpResponse with CSV file
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    writer = csv.writer(response)
    
    # Write header
    writer.writerow([display_name for _, display_name in fields])
    
    # Write data rows
    for obj in queryset:
        row = []
        for field_name, _ in fields:
            # Handle nested attributes (e.g., 'product.name')
            value = obj
            for attr in field_name.split('.'):
                value = getattr(value, attr, '')
            row.append(str(value) if value is not None else '')
        writer.writerow(row)
    
    return response


def export_to_excel(queryset, fields, filename='export', sheet_name='Data'):
    """
    Export a queryset to Excel format with styling.
    
    Args:
        queryset: Django queryset to export
        fields: List of tuples (field_name, display_name)
        filename: Base filename for the export
        sheet_name: Name of the Excel sheet
    
    Returns:
        HttpResponse with Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Header styling
    header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write header
    for col_num, (_, display_name) in enumerate(fields, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = display_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Write data rows
    for row_num, obj in enumerate(queryset, 2):
        for col_num, (field_name, _) in enumerate(fields, 1):
            # Handle nested attributes
            value = obj
            for attr in field_name.split('.'):
                value = getattr(value, attr, '')
            
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = str(value) if value is not None else ''
            cell.alignment = Alignment(vertical='center')
    
    # Auto-adjust column widths
    for col_num in range(1, len(fields) + 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response


def export_sales_report(orders, format='csv'):
    """Export sales report with order details."""
    fields = [
        ('order_number', 'Order Number'),
        ('customer.name', 'Customer'),
        ('created_at', 'Date'),
        ('total_amount', 'Total Amount'),
        ('payment_method', 'Payment Method'),
        ('status', 'Status'),
        ('branch.name', 'Branch'),
    ]
    
    filename = 'sales_report'
    
    if format == 'excel':
        return export_to_excel(orders, fields, filename, 'Sales Report')
    else:
        return export_to_csv(orders, fields, filename)


def export_inventory_report(products, format='csv'):
    """Export inventory report with product details."""
    fields = [
        ('sku', 'SKU'),
        ('name', 'Product Name'),
        ('category.name', 'Category'),
        ('stock_quantity', 'Stock Quantity'),
        ('price', 'Price'),
        ('cost', 'Cost'),
        ('branch.name', 'Branch'),
        ('is_active', 'Active'),
    ]
    
    filename = 'inventory_report'
    
    if format == 'excel':
        return export_to_excel(products, fields, filename, 'Inventory')
    else:
        return export_to_csv(products, fields, filename)


def export_customer_report(customers, format='csv'):
    """Export customer report with contact details."""
    fields = [
        ('name', 'Name'),
        ('email', 'Email'),
        ('phone', 'Phone'),
        ('loyalty_points', 'Loyalty Points'),
        ('total_purchases', 'Total Purchases'),
        ('created_at', 'Member Since'),
    ]
    
    filename = 'customer_report'
    
    if format == 'excel':
        return export_to_excel(customers, fields, filename, 'Customers')
    else:
        return export_to_csv(customers, fields, filename)


def export_order_items_report(order_items, format='csv'):
    """Export detailed order items report."""
    fields = [
        ('order.order_number', 'Order Number'),
        ('product.name', 'Product'),
        ('quantity', 'Quantity'),
        ('price', 'Unit Price'),
        ('subtotal', 'Subtotal'),
        ('order.created_at', 'Date'),
    ]
    
    filename = 'order_items_report'
    
    if format == 'excel':
        return export_to_excel(order_items, fields, filename, 'Order Items')
    else:
        return export_to_csv(order_items, fields, filename)
